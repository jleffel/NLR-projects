"""
AHA Hospital Dataset Analysis
Classifies hospitals into rural/nominal/campus categories using the following priority:
  1. Sq ft available → classify by sq ft (tiebreaker for edge cases)
  2. Sq ft missing, numeric bed count available → classify by bed count
  3. Both missing → classify by Bed size code midpoint

Edge cases (sq ft and bed count disagree) are resolved via sq ft tiebreaker,
folded back into Categories with Edge_Case_Flag=True, and also listed in EdgeCases
for reference.

Output: Multi-sheet Excel workbook with:
  - Categories  : All classified hospitals (including resolved edge cases)
  - Stats        : IQR, +/-1.5 SD, min/max ratio — all classified hospitals
  - Stats_Clean  : Same stats restricted to SqFt+Beds (agree) hospitals only (no edge cases, no imputed)
  - EdgeCases    : Hospitals where sq ft and bed count conflicted (resolved via sq ft tiebreaker)
  - Outliers     : Hospitals outside all three category ranges on all available dimensions
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── File paths ────────────────────────────────────────────────────────────────
# Where to read from and where to save the output
INPUT_FILE  = r"C:\Users\Jessica\Desktop\NLR\Hospital Project\AHA Annotated Hospital Data.xlsx"
OUTPUT_FILE = r"C:\Users\Jessica\Desktop\NLR\Hospital Project\AHA_Hospital_Analysis_Output.xlsx"
INPUT_SHEET = "CY"

# ── Category thresholds (Table 1) ─────────────────────────────────────────────
# The sq ft and bed count ranges that define each hospital category
THRESHOLDS = {
    "Rural":   {"sqft": (1_215,     150_000), "beds": (2,    250)},
    "Nominal": {"sqft": (150_001,   860_000), "beds": (20,   650)},
    "Campus":  {"sqft": (860_001, 5_000_000), "beds": (100, 1800)},
}

CATEGORIES = ["Rural", "Nominal", "Campus"]

# ── Bed size code → midpoint ───────────────────────────────────────────────────
# Converts the text bed size range (e.g. "6-24 beds") into a single number
# so we can classify hospitals that are missing both sq ft and numeric bed count.
# "500 or more beds" uses 500 as a conservative floor (maps to Campus).
BED_SIZE_CODE_MIDPOINT = {
    "6-24 beds":        15,
    "25-49 beds":       37,
    "50-99 beds":       75,
    "100-199 beds":    150,
    "200-299 beds":    250,
    "300-399 beds":    350,
    "400-499 beds":    450,
    "500 or more beds": 500,
}

# ── Column names ──────────────────────────────────────────────────────────────
# These must match the exact header names in the AHA Excel file
COL_ID       = "AHA ID"
COL_SQFT     = "Total gross square feet"
COL_BEDS     = "Total hospital beds"
COL_BEDCODE  = "Bed size code"
COL_TEACHING = "Teaching Status"
COL_FOC      = "Freestanding outpatient center"


# ── Classification helper functions ───────────────────────────────────────────

def classify_by_sqft(sqft):
    # Look up which category a sq ft value falls into
    # Returns None if the value is missing or outside all ranges
    if pd.isna(sqft):
        return None
    for cat, bounds in THRESHOLDS.items():
        if bounds["sqft"][0] <= sqft <= bounds["sqft"][1]:
            return cat
    return None


def classify_by_beds(beds):
    # Look up which category a bed count falls into
    # Returns None if the value is missing or outside all ranges
    if pd.isna(beds):
        return None
    for cat, bounds in THRESHOLDS.items():
        if bounds["beds"][0] <= beds <= bounds["beds"][1]:
            return cat
    return None


def classify_by_bed_code(code):
    # Convert the text bed size code to a midpoint number,
    # then classify using the same bed count ranges as classify_by_beds()
    if pd.isna(code):
        return None
    midpoint = BED_SIZE_CODE_MIDPOINT.get(str(code).strip())
    return classify_by_beds(midpoint) if midpoint is not None else None


# ── Statistics function ────────────────────────────────────────────────────────

def compute_stats(series, label):
    # Given a list of sq ft per bed values and a group label,
    # compute IQR, mean, std, +/-1.5 SD bounds, and min/max within those bounds
    # Returns a row of results as a dictionary
    n = len(series)
    if n == 0:
        # If no hospitals in this group, return a row of blank values
        return {
            "Group": label, "N": 0,
            "Q1": np.nan, "Median": np.nan, "Q3": np.nan, "IQR": np.nan,
            "Mean": np.nan, "Std": np.nan,
            "Lower_1.5SD": np.nan, "Upper_1.5SD": np.nan,
            "Min_in_bounds": np.nan, "Max_in_bounds": np.nan,
        }
    # Calculate quartiles and spread
    q1  = series.quantile(0.25)
    med = series.quantile(0.50)
    q3  = series.quantile(0.75)
    iqr = q3 - q1
    mu  = series.mean()
    sd  = series.std(ddof=1)
    # Calculate the acceptable range as mean +/- 1.5 standard deviations
    lower  = mu - 1.5 * sd
    upper  = mu + 1.5 * sd
    # Find actual min and max values that fall within that range
    within = series[(series >= lower) & (series <= upper)]
    return {
        "Group":         label,
        "N":             n,
        "Q1":            round(q1,  2),
        "Median":        round(med, 2),
        "Q3":            round(q3,  2),
        "IQR":           round(iqr, 2),
        "Mean":          round(mu,  2),
        "Std":           round(sd,  2),
        "Lower_1.5SD":   round(lower, 2),
        "Upper_1.5SD":   round(upper, 2),
        "Min_in_bounds": round(within.min(), 2) if len(within) else np.nan,
        "Max_in_bounds": round(within.max(), 2) if len(within) else np.nan,
    }


# ── Summary table builder ──────────────────────────────────────────────────────

def build_summary_rows(df_classified):
    # For each category (Rural/Nominal/Campus), compute stats for:
    #   1. All hospitals in that category
    #   2. Each teaching status sub-group within that category
    #   3. Each FOC (freestanding outpatient center) sub-group within that category
    rows = []
    for cat in CATEGORIES:
        # All hospitals in this category
        sub = df_classified[df_classified["Category"] == cat]["SqFt_Per_Bed"].dropna()
        rows.append(compute_stats(sub, f"{cat} (All)"))

        # Break down by teaching status (e.g. Major, Minor, Non Teaching)
        for tv in sorted(df_classified[df_classified["Category"] == cat][COL_TEACHING].dropna().unique()):
            mask = (df_classified["Category"] == cat) & (df_classified[COL_TEACHING] == tv)
            rows.append(compute_stats(df_classified[mask]["SqFt_Per_Bed"].dropna(),
                                      f"  {cat} | Teaching={tv}"))

        # Break down by FOC presence (Yes/No)
        for fv in sorted(df_classified[df_classified["Category"] == cat][COL_FOC].dropna().unique()):
            mask = (df_classified["Category"] == cat) & (df_classified[COL_FOC] == fv)
            rows.append(compute_stats(df_classified[mask]["SqFt_Per_Bed"].dropna(),
                                      f"  {cat} | FOC={fv}"))
    return pd.DataFrame(rows)


# ── Excel styling function ─────────────────────────────────────────────────────

def style_stats_sheet(ws):
    # Make the header row dark blue with white bold text
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Highlight primary category rows (Rural/Nominal/Campus All) in light blue
    # Sub-group rows (teaching/FOC breakdowns) get plain formatting
    cat_fill = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        is_primary = row[0].value and not str(row[0].value).startswith(" ")
        for cell in row:
            if is_primary:
                cell.fill = cat_fill
                cell.font = Font(bold=True, name="Arial")
            else:
                cell.font = Font(name="Arial")

    # Auto-size each column to fit its content
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


# ── Main script ────────────────────────────────────────────────────────────────

def main():

    # STEP 1: Load the AHA Excel file and keep only the columns we need
    print(f"Reading {INPUT_FILE} ...")
    df = pd.read_excel(INPUT_FILE, sheet_name=INPUT_SHEET, dtype={COL_ID: str})

    # Check that all required columns exist; raise an error if any are missing
    needed = [COL_ID, COL_SQFT, COL_BEDS, COL_BEDCODE, COL_TEACHING, COL_FOC]
    missing_cols = [c for c in needed if c not in df.columns]
    if missing_cols:
        raise KeyError(
            f"Column(s) not found in sheet '{INPUT_SHEET}': {missing_cols}\n"
            f"Available columns: {list(df.columns)}"
        )

    # Keep only the columns we need and convert sq ft / beds to numbers
    df = df[needed].copy()
    df[COL_SQFT] = pd.to_numeric(df[COL_SQFT], errors="coerce")
    df[COL_BEDS] = pd.to_numeric(df[COL_BEDS], errors="coerce")

    # Drop blank rows at the bottom of the sheet (no AHA ID means not a real hospital)
    before = len(df)
    df = df[df[COL_ID].notna()].copy()
    dropped_blank = before - len(df)
    if dropped_blank:
        print(f"Dropped {dropped_blank} blank/junk rows (no AHA ID).")

    # STEP 2: Pre-classify each hospital using all three available methods
    # We compute all three up front so we can compare them in the loop below
    df["Cat_SqFt"]    = df[COL_SQFT].apply(classify_by_sqft)     # category from sq ft
    df["Cat_Beds"]    = df[COL_BEDS].apply(classify_by_beds)      # category from bed count
    df["Cat_BedCode"] = df[COL_BEDCODE].apply(classify_by_bed_code)  # category from bed size code text

    # STEP 3: Loop through every hospital and decide its final category
    # Hospitals go into one of three lists depending on the outcome
    records_classified = []   # successfully categorized hospitals
    records_edge       = []   # hospitals where sq ft and beds disagreed (kept as reference)
    records_outlier    = []   # hospitals that couldn't be categorized at all

    for _, row in df.iterrows():
        # Check what data is available for this hospital
        has_sqft    = pd.notna(row[COL_SQFT])
        has_beds    = pd.notna(row[COL_BEDS])
        cat_sqft    = row["Cat_SqFt"]
        cat_beds    = row["Cat_Beds"]
        cat_bedcode = row["Cat_BedCode"]
        # Start building the output row without the temporary classification columns
        r = row.drop(["Cat_SqFt", "Cat_Beds", "Cat_BedCode"]).to_dict()

        if has_sqft and has_beds:
            # Both sq ft and bed count are available — compare their categories
            if cat_sqft == cat_beds:
                # They agree: use that category (or flag as outlier if both are out of range)
                if cat_sqft is not None:
                    r["Category"] = cat_sqft
                    r["Classification_Method"] = "SqFt+Beds (agree)"
                    r["Edge_Case_Flag"] = False
                    records_classified.append(r)
                else:
                    # Both values exist but neither falls in any category range
                    r["Outlier_Reason"] = "SqFt and Beds both outside all category ranges"
                    records_outlier.append(r)
            else:
                # They disagree: apply sq ft as the tiebreaker
                # Save both conflicting categories for reference in the EdgeCases sheet
                r["Cat_SqFt_Flag"] = cat_sqft
                r["Cat_Beds_Flag"] = cat_beds
                if cat_sqft is not None:
                    # Sq ft gives a valid category — use it
                    r["Category"] = cat_sqft
                    r["Classification_Method"] = "SqFt tiebreaker"
                elif cat_beds is not None:
                    # Sq ft is out of range but beds gives a valid category — use beds as fallback
                    r["Category"] = cat_beds
                    r["Classification_Method"] = "Beds fallback (SqFt out of range)"
                else:
                    # Neither sq ft nor beds falls in any range — unresolvable
                    r["Category"] = None
                    r["Classification_Method"] = "Unresolvable"

                if r["Category"] is not None:
                    # Add to classified (without the flag columns) and also to EdgeCases for reference
                    classified_r = {k: v for k, v in r.items()
                                    if k not in ["Cat_SqFt_Flag", "Cat_Beds_Flag"]}
                    classified_r["Edge_Case_Flag"] = True
                    records_classified.append(classified_r)
                    records_edge.append(r)   # full record with both flags goes to EdgeCases sheet
                else:
                    r["Outlier_Reason"] = "SqFt and Beds both outside all ranges (conflict)"
                    records_outlier.append(r)

        elif has_sqft:
            # Only sq ft is available — use it alone
            if cat_sqft is not None:
                r["Category"] = cat_sqft
                r["Classification_Method"] = "SqFt only (Beds missing)"
                r["Edge_Case_Flag"] = False
                records_classified.append(r)
            else:
                r["Outlier_Reason"] = "SqFt outside all category ranges; Beds missing"
                records_outlier.append(r)

        elif has_beds:
            # Only bed count is available — use it alone
            if cat_beds is not None:
                r["Category"] = cat_beds
                r["Classification_Method"] = "Beds only (SqFt missing)"
                r["Edge_Case_Flag"] = False
                records_classified.append(r)
            else:
                r["Outlier_Reason"] = "Beds outside all category ranges; SqFt missing"
                records_outlier.append(r)

        else:
            # Neither sq ft nor numeric beds available — fall back to bed size code text
            if cat_bedcode is not None:
                r["Category"] = cat_bedcode
                r["Classification_Method"] = "Bed size code (SqFt+Beds missing)"
                r["Edge_Case_Flag"] = False
                records_classified.append(r)
            elif pd.notna(row[COL_BEDCODE]):
                # Bed size code exists but maps outside all ranges
                r["Outlier_Reason"] = "Bed size code outside all category ranges"
                records_outlier.append(r)
            else:
                # Absolutely no usable data — cannot classify
                r["Outlier_Reason"] = "SqFt, Beds, and Bed size code all missing"
                records_outlier.append(r)

    # STEP 4: Convert the three lists into dataframes
    df_classified = pd.DataFrame(records_classified)
    df_edges      = pd.DataFrame(records_edge)
    df_outliers   = pd.DataFrame(records_outlier)

    # Calculate sq ft per bed ratio — only possible when both sq ft and numeric beds exist
    # Hospitals classified by bed count alone will have NaN here, which is correct
    for frame in [df_classified, df_edges]:
        if len(frame):
            frame["SqFt_Per_Bed"] = np.where(
                frame[COL_BEDS].notna() & (frame[COL_BEDS] > 0) & frame[COL_SQFT].notna(),
                frame[COL_SQFT] / frame[COL_BEDS],
                np.nan,
            )

    # STEP 5: Compute summary statistics for all classified hospitals
    df_summary = build_summary_rows(df_classified)

    # Also compute stats for the cleanest subset only:
    # hospitals where sq ft and beds both existed and agreed (no imputed, no edge cases)
    df_clean = df_classified[df_classified["Classification_Method"] == "SqFt+Beds (agree)"].copy()
    df_summary_clean = build_summary_rows(df_clean)

    # STEP 6: Print summary to console
    print("\n=== Primary Category Summary — All Classified (SqFt per Bed) ===")
    primary_rows = df_summary[~df_summary["Group"].str.startswith(" ")]
    print(primary_rows.to_string(index=False))

    print("\n=== Primary Category Summary — Clean Only (SqFt+Beds agree) ===")
    primary_clean = df_summary_clean[~df_summary_clean["Group"].str.startswith(" ")]
    print(primary_clean.to_string(index=False))

    print(f"\nTotal classified (incl. resolved edge cases) : {len(df_classified)}")
    print(f"Edge cases (conflict flagged, also in above)  : {len(df_edges)}")
    print(f"Outliers                                      : {len(df_outliers)}")
    print(f"Grand total                                   : {len(df_classified) + len(df_outliers)}")

    print("\n=== Classification Method Breakdown ===")
    print(df_classified["Classification_Method"].value_counts().to_string())

    print("\n=== Category Counts (all classified) ===")
    print(df_classified["Category"].value_counts().to_string())

    # STEP 7: Reorder columns so the most important ones appear first in the output
    front_cols = [COL_ID, COL_SQFT, COL_BEDS, COL_BEDCODE,
                  COL_TEACHING, COL_FOC,
                  "Category", "Classification_Method", "Edge_Case_Flag", "SqFt_Per_Bed"]
    front_cols = [c for c in front_cols if c in df_classified.columns]
    remaining  = [c for c in df_classified.columns if c not in front_cols]
    df_classified = df_classified[front_cols + remaining]

    # STEP 8: Write all results to a multi-sheet Excel workbook
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df_classified.to_excel(writer,       sheet_name="Categories",  index=False)
        df_summary.to_excel(writer,          sheet_name="Stats",       index=False)
        df_summary_clean.to_excel(writer,    sheet_name="Stats_Clean", index=False)
        df_edges.to_excel(writer,            sheet_name="EdgeCases",   index=False)
        df_outliers.to_excel(writer,         sheet_name="Outliers",    index=False)

    # Apply formatting to the two stats sheets
    wb = load_workbook(OUTPUT_FILE)
    style_stats_sheet(wb["Stats"])
    style_stats_sheet(wb["Stats_Clean"])
    wb.save(OUTPUT_FILE)

    print(f"\nOutput written to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
