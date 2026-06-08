"""
AHA Hospital Dataset Analysis
Classifies hospitals into rural/nominal/campus categories based on square footage
and bed count thresholds, computes IQR and +/-1.5 SD statistics on sq ft per bed ratio,
and provides secondary breakdowns by teaching status and freestanding outpatient center (FOC).

Output: Multi-sheet Excel workbook with:
  - CY_Classified   : Full dataset with assigned category
  - Summary         : IQR, +/-1.5 SD, min/max ratio per category + teaching/FOC breakdowns
  - Edge_Cases      : Hospitals where sq ft and bed count point to different categories
  - Outliers        : Hospitals outside all three category ranges
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── File paths ────────────────────────────────────────────────────────────────
INPUT_FILE  = r"C:\Users\jleffel\OneDrive - NLR\Hospital Project\AHA Annotated Hospital Data.xlsx"
OUTPUT_FILE = r"C:\Users\jleffel\OneDrive - NLR\Hospital Project\AHA_Hospital_Analysis_Output.xlsx"
INPUT_SHEET = "CY"

# ── Category thresholds (Table 1) ─────────────────────────────────────────────
THRESHOLDS = {
    "Rural":   {"sqft": (1_215,     150_000),   "beds": (2,    250)},
    "Nominal": {"sqft": (150_001,   860_000),   "beds": (20,   650)},
    "Campus":  {"sqft": (860_001, 5_000_000),   "beds": (100, 1800)},
}

CATEGORIES = ["Rural", "Nominal", "Campus"]

# ── Column names (adjust if actual headers differ) ────────────────────────────
COL_ID       = "AHA ID"
COL_SQFT     = "Total gross square feet"
COL_BEDS     = "Total hospital beds"
COL_TEACHING = "Teaching Status"
COL_FOC      = "Freestanding outpatient center"


def classify_by_sqft(sqft):
    """Return category name based solely on sq ft range, or None."""
    for cat, bounds in THRESHOLDS.items():
        lo, hi = bounds["sqft"]
        if lo <= sqft <= hi:
            return cat
    return None


def classify_by_beds(beds):
    """Return category name based solely on bed count range, or None."""
    for cat, bounds in THRESHOLDS.items():
        lo, hi = bounds["beds"]
        if lo <= beds <= hi:
            return cat
    return None


def compute_stats(series, label):
    """
    Given a numeric Series (sq ft / bed ratio), return a dict of stats.
    IQR, mean +/- 1.5 SD bounds, and the min/max values within those bounds.
    """
    n = len(series)
    if n == 0:
        return {
            "Group": label, "N": 0,
            "Q1": np.nan, "Median": np.nan, "Q3": np.nan, "IQR": np.nan,
            "Mean": np.nan, "Std": np.nan,
            "Lower_1.5SD": np.nan, "Upper_1.5SD": np.nan,
            "Min_in_bounds": np.nan, "Max_in_bounds": np.nan,
        }

    q1  = series.quantile(0.25)
    med = series.quantile(0.50)
    q3  = series.quantile(0.75)
    iqr = q3 - q1
    mu  = series.mean()
    sd  = series.std(ddof=1)

    lower = mu - 1.5 * sd
    upper = mu + 1.5 * sd

    within = series[(series >= lower) & (series <= upper)]

    return {
        "Group":        label,
        "N":            n,
        "Q1":           round(q1,  2),
        "Median":       round(med, 2),
        "Q3":           round(q3,  2),
        "IQR":          round(iqr, 2),
        "Mean":         round(mu,  2),
        "Std":          round(sd,  2),
        "Lower_1.5SD":  round(lower, 2),
        "Upper_1.5SD":  round(upper, 2),
        "Min_in_bounds": round(within.min(), 2) if len(within) else np.nan,
        "Max_in_bounds": round(within.max(), 2) if len(within) else np.nan,
    }


def build_summary_rows(df_classified):
    """
    Build all summary stat rows:
      - One row per primary category
      - Within each category, one row per Teaching Status value
      - Within each category, one row per FOC value
    """
    rows = []

    for cat in CATEGORIES:
        sub = df_classified[df_classified["Category"] == cat]["SqFt_Per_Bed"].dropna()
        rows.append(compute_stats(sub, f"{cat} (All)"))

        # Teaching status breakdown
        teaching_vals = (
            df_classified[df_classified["Category"] == cat][COL_TEACHING]
            .dropna()
            .unique()
        )
        for tv in sorted(teaching_vals):
            mask = (df_classified["Category"] == cat) & (df_classified[COL_TEACHING] == tv)
            sub_t = df_classified[mask]["SqFt_Per_Bed"].dropna()
            rows.append(compute_stats(sub_t, f"  {cat} | Teaching={tv}"))

        # FOC breakdown
        foc_vals = (
            df_classified[df_classified["Category"] == cat][COL_FOC]
            .dropna()
            .unique()
        )
        for fv in sorted(foc_vals):
            mask = (df_classified["Category"] == cat) & (df_classified[COL_FOC] == fv)
            sub_f = df_classified[mask]["SqFt_Per_Bed"].dropna()
            rows.append(compute_stats(sub_f, f"  {cat} | FOC={fv}"))

    return pd.DataFrame(rows)


def style_summary_sheet(ws):
    """Apply basic header formatting to the Summary sheet."""
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")

    for cell in ws[1]:
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Highlight primary category rows (those that don't start with spaces)
    cat_fill = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value and not str(row[0].value).startswith(" "):
            for cell in row:
                cell.fill = cat_fill
                cell.font = Font(bold=True, name="Arial")
        else:
            for cell in row:
                cell.font = Font(name="Arial")

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def main():
    # ── Load data ─────────────────────────────────────────────────────────────
    print(f"Reading {INPUT_FILE} ...")
    df = pd.read_excel(
        INPUT_FILE,
        sheet_name=INPUT_SHEET,
        dtype={COL_ID: str},
    )

    needed = [COL_ID, COL_SQFT, COL_BEDS, COL_TEACHING, COL_FOC]
    missing_cols = [c for c in needed if c not in df.columns]
    if missing_cols:
        raise KeyError(
            f"Column(s) not found in sheet '{INPUT_SHEET}': {missing_cols}\n"
            f"Available columns: {list(df.columns)}"
        )

    df = df[needed].copy()

    # ── Drop rows missing both key numeric fields ──────────────────────────────
    df_work = df.dropna(subset=[COL_SQFT, COL_BEDS]).copy()
    df_work[COL_SQFT] = pd.to_numeric(df_work[COL_SQFT], errors="coerce")
    df_work[COL_BEDS] = pd.to_numeric(df_work[COL_BEDS], errors="coerce")
    df_work = df_work.dropna(subset=[COL_SQFT, COL_BEDS])

    # ── Classify ──────────────────────────────────────────────────────────────
    df_work["Cat_SqFt"] = df_work[COL_SQFT].apply(classify_by_sqft)
    df_work["Cat_Beds"] = df_work[COL_BEDS].apply(classify_by_beds)

    # Outliers: outside all ranges on both dimensions
    outliers_mask = df_work["Cat_SqFt"].isna() & df_work["Cat_Beds"].isna()
    df_outliers   = df_work[outliers_mask].drop(columns=["Cat_SqFt", "Cat_Beds"])

    # Edge cases: sq ft and bed count disagree (and at least one is non-null)
    df_valid = df_work[~outliers_mask].copy()
    edge_mask   = df_valid["Cat_SqFt"] != df_valid["Cat_Beds"]
    df_edges    = df_valid[edge_mask].copy()
    df_edges["Cat_SqFt_Flag"] = df_edges["Cat_SqFt"]
    df_edges["Cat_Beds_Flag"] = df_edges["Cat_Beds"]
    df_edges = df_edges.drop(columns=["Cat_SqFt", "Cat_Beds"])

    # Classified: sq ft and beds agree
    df_classified = df_valid[~edge_mask].copy()
    df_classified["Category"] = df_classified["Cat_SqFt"]
    df_classified = df_classified.drop(columns=["Cat_SqFt", "Cat_Beds"])

    # Sq ft per bed ratio (guard against zero beds)
    df_classified["SqFt_Per_Bed"] = np.where(
        df_classified[COL_BEDS] > 0,
        df_classified[COL_SQFT] / df_classified[COL_BEDS],
        np.nan,
    )

    # ── Summary stats ─────────────────────────────────────────────────────────
    df_summary = build_summary_rows(df_classified)

    # ── Console print of primary category stats ────────────────────────────────
    print("\n=== Primary Category Summary (SqFt per Bed) ===")
    primary_rows = df_summary[~df_summary["Group"].str.startswith(" ")]
    print(primary_rows.to_string(index=False))
    print(f"\nTotal classified : {len(df_classified)}")
    print(f"Edge cases       : {len(df_edges)}")
    print(f"Outliers         : {len(df_outliers)}")

    # ── Write output workbook ─────────────────────────────────────────────────
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df_classified.drop(columns=["SqFt_Per_Bed"]).to_excel(
            writer, sheet_name="CY_Classified", index=False
        )
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_edges.to_excel(writer, sheet_name="Edge_Cases", index=False)
        df_outliers.to_excel(writer, sheet_name="Outliers", index=False)

    # Apply styling to Summary sheet
    wb = load_workbook(OUTPUT_FILE)
    style_summary_sheet(wb["Summary"])
    wb.save(OUTPUT_FILE)

    print(f"\nOutput written to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
